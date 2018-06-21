import openpyxl
import requests
from bs4 import BeautifulSoup

response = requests.get('http://search.11st.co.kr/Search.tmall?kwd=%25EB%259D%25BC%25EC%25A6%2588%25EB%25B2%25A0%25EB%25A6%25AC%25ED%258C%258C%25EC%259D%25B4')

soup = BeautifulSoup(response.content, 'html.parser')

first = soup.find("ul", {"data-log-actionid-area":"hot"})
twice = first.find("p", {"class":"info_tit"})

print(first.get_text())
print(twice.get_text())

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Hello'
ws['A2'] = 'Hello2'
wb.save('test.xlsx')

# for title in first.find_all('a'):
#     print(title.get_text())


# final = first.find("a")

# print(first.get_text())

